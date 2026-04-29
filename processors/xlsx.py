"""XLSX → Markdown：用 openpyxl 提取。"""
from pathlib import Path
from openpyxl import load_workbook


def process_xlsx(xlsx_path: str, output_path: str, config: dict | None = None) -> dict:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    parts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if len(parts) > 0:
            parts.append("")
        parts.append(f"## {sheet_name}")

        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                parts.append(" | ".join(cells))

    wb.close()
    text = "\n".join(parts)
    Path(output_path).write_text(text, "utf-8", newline="\r\n")
    return {"status": "ok", "chars": len(text)}
